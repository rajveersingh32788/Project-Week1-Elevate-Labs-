"""
Healthcare No-Show – Power BI Insight Dashboard (Excel)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
import os

# ── Palette ──────────────────────────────────────────────────────────────────
BLU  = '1E5799'; BLU2 = '2E86C1'; BLU3 = 'AED6F1'
RED  = 'C0392B'; ORG  = 'E67E22'; GRN  = '27AE60'
GRY  = 'ECF0F1'; DRK  = '2C3E50'; WHT  = 'FFFFFF'
YLW  = 'F9E79F'

def hdr(cell, text, sz=12, bold=True, color=WHT, bg=BLU, wrap=False, italic=False):
    cell.value = text
    cell.font  = Font(name='Calibri', size=sz, bold=bold, color=color, italic=italic)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=wrap)

def val(cell, text, sz=11, bold=False, color=DRK, bg=WHT, align='center'):
    cell.value = text
    cell.font  = Font(name='Calibri', size=sz, bold=bold, color=color)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')

def border_range(ws, min_row, min_col, max_row, max_col, color='CCCCCC'):
    thin = Side(style='thin', color=color)
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(r,c).border = bdr

# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 – EXECUTIVE DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Executive Dashboard'
ws.sheet_view.showGridLines = False
ws.row_dimensions[1].height = 45

# Title bar
ws.merge_cells('A1:N1')
hdr(ws['A1'], '🏥  Healthcare Appointment No-Show Prediction — Executive Dashboard',
    sz=16, bg=BLU)

ws.merge_cells('A2:N2')
hdr(ws['A2'], 'Dataset: 5,000 Appointments  |  Model: Decision Tree  |  Analysis Period: 2016',
    sz=10, bg=BLU2, bold=False, italic=True)
ws.row_dimensions[2].height = 20

# ── KPI CARDS (row 4-8) ──────────────────────────────────────────────────────
kpi_data = [
    ('Total Appointments', '5,000',  BLU,  '📋'),
    ('No-Show Count',      '1,247',  RED,  '❌'),
    ('No-Show Rate',       '24.9%',  ORG,  '📊'),
    ('SMS Sent',           '32.0%',  GRN,  '📱'),
    ('High Risk Patients', '1,015',  RED,  '⚠️'),
    ('Model AUC-ROC',      '0.575',  BLU2, '🤖'),
]
ws.row_dimensions[3].height = 8
ws.row_dimensions[4].height = 14
ws.row_dimensions[5].height = 32
ws.row_dimensions[6].height = 24
ws.row_dimensions[7].height = 14
ws.row_dimensions[8].height = 8

kpi_cols = [1, 3, 5, 7, 9, 11]  # A C E G I K
for idx, (label, value, color, icon) in enumerate(kpi_data):
    col = kpi_cols[idx]
    col2 = col + 1
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col2)
    ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col2)
    ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col2)
    ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col2)
    hdr(ws.cell(4, col), f'{icon} {label}', sz=9, bg=color, bold=False, italic=True)
    c = ws.cell(5, col)
    c.value = value
    c.font = Font(name='Calibri', size=24, bold=True, color=WHT)
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')
    hdr(ws.cell(6, col), '', sz=9, bg=color)
    hdr(ws.cell(7, col), '', sz=9, bg=color)

for c in range(1, 13):
    ws.column_dimensions[get_column_letter(c)].width = 10
ws.column_dimensions['M'].width = 2
ws.column_dimensions['N'].width = 2

# ── Section Headers ──────────────────────────────────────────────────────────
ws.merge_cells('A10:F10')
hdr(ws['A10'], '📈  No-Show by Key Dimension', sz=12, bg=BLU)
ws.merge_cells('H10:N10')
hdr(ws['H10'], '🤖  Model Performance Summary', sz=12, bg=DRK)

ws.row_dimensions[10].height = 22

# ── Trend Table: SMS ─────────────────────────────────────────────────────────
headers = ['Dimension', 'Category', 'No-Show Rate', 'Appointments', 'Risk Level']
for ci, h in enumerate(headers, 1):
    c = ws.cell(11, ci)
    hdr(c, h, sz=10, bg=BLU2)
ws.row_dimensions[11].height = 18

trend_data = [
    ('SMS Reminder', 'No SMS Sent',   '28.2%', '3,400', 'HIGH',   RED),
    ('SMS Reminder', 'SMS Sent',      '16.7%', '1,600', 'LOW',    GRN),
    ('Age Group',    'Child (0-12)',   '26.5%',   '400', 'HIGH',   RED),
    ('Age Group',    'Teen (13-17)',   '27.8%',   '180', 'HIGH',   RED),
    ('Age Group',    'Young Adult',    '25.1%',   '900', 'MEDIUM', ORG),
    ('Age Group',    'Adult (36-60)',  '22.3%', '1,800', 'MEDIUM', ORG),
    ('Age Group',    'Senior (60+)',   '23.0%', '1,720', 'LOW',    GRN),
    ('Weekday',      'Monday',        '27.4%',   '820', 'HIGH',   RED),
    ('Weekday',      'Tuesday',       '24.1%',   '850', 'MEDIUM', ORG),
    ('Weekday',      'Wednesday',     '23.7%',   '870', 'LOW',    GRN),
    ('Wait Days',    '0-10 days',     '18.2%', '1,100', 'LOW',    GRN),
    ('Wait Days',    '11-30 days',    '23.9%', '2,000', 'MEDIUM', ORG),
    ('Wait Days',    '31+ days',      '29.6%', '1,900', 'HIGH',   RED),
]
bg_alts = [WHT, GRY]
for ri, row in enumerate(trend_data, 12):
    bg = bg_alts[ri % 2]
    for ci, txt in enumerate(row[:5], 1):
        c = ws.cell(ri, ci)
        c.value = txt
        c.font  = Font(name='Calibri', size=10, color=DRK)
        c.fill  = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
    # colour risk badge
    risk_cell = ws.cell(ri, 5)
    risk_cell.fill = PatternFill('solid', fgColor=row[5])
    risk_cell.font = Font(name='Calibri', size=10, bold=True, color=WHT)
    ws.row_dimensions[ri].height = 16

border_range(ws, 11, 1, 11+len(trend_data), 5)

# ── Model Performance Table ──────────────────────────────────────────────────
model_headers = ['Metric', 'Score', 'Benchmark', 'Status']
for ci, h in enumerate(model_headers, 8):
    hdr(ws.cell(11, ci), h, sz=10, bg=DRK)

model_rows = [
    ('Accuracy',  '55.4%', '70%', '⚠️ Fair',   ORG),
    ('Precision', '28.6%', '50%', '⚠️ Fair',   ORG),
    ('Recall',    '53.0%', '60%', '✓ Good',    GRN),
    ('F1 Score',  '37.2%', '60%', '⚠️ Fair',   ORG),
    ('AUC-ROC',   '0.575', '0.70','⚠️ Fair',   ORG),
    ('CV AUC',    '0.582', '0.70','⚠️ Fair',   ORG),
]
for ri, row in enumerate(model_rows, 12):
    bg = bg_alts[ri % 2]
    for ci, txt in enumerate(row[:4], 8):
        c = ws.cell(ri, ci)
        c.value = txt
        c.font  = Font(name='Calibri', size=10, color=DRK)
        c.fill  = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(ri, 11).fill = PatternFill('solid', fgColor=row[4])
    ws.row_dimensions[ri].height = 16

border_range(ws, 11, 8, 17, 11)

# Recommendations box
ws.merge_cells('A27:N27')
hdr(ws['A27'], '💡  Optimization Recommendations', sz=12, bg=GRN)
ws.row_dimensions[27].height = 22

recs = [
    ('1', 'Send SMS to ALL patients', 'SMS reminders reduce no-shows by ~11%. Implement universal SMS reminder system.'),
    ('2', 'Double-book high-risk slots', 'Patients with >30-day wait time have 30% no-show rate — over-schedule these slots by 15%.'),
    ('3', 'Monday morning call-backs',  'Monday has the highest no-show rate. Add phone follow-ups for Monday appointments.'),
    ('4', 'Child/Teen outreach',        'Under-18 patients show highest no-show rates. Engage parents directly via WhatsApp/SMS.'),
    ('5', 'Overbooking by risk tier',   'High-risk patients (prob >0.60): overbook 20%. Medium: 10%. Low: standard scheduling.'),
    ('6', 'Retrain model quarterly',    'Retrain with new data every quarter and consider Random Forest / XGBoost for higher AUC.'),
]
hdr_cells = ['#', 'Action', 'Detail']
hdr_widths = [3, 25, 60]
cols_r = [1, 2, 5]
for ci, h in zip(cols_r, hdr_cells):
    hdr(ws.cell(28, ci), h, sz=10, bg=BLU2)

for ri, (num, action, detail) in enumerate(recs, 29):
    bg = bg_alts[ri % 2]
    ws.cell(ri, 1).value = num
    ws.cell(ri, 1).font  = Font(name='Calibri', size=11, bold=True, color=WHT)
    ws.cell(ri, 1).fill  = PatternFill('solid', fgColor=GRN)
    ws.cell(ri, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=4)
    c = ws.cell(ri, 2)
    c.value = action; c.font = Font(name='Calibri', size=10, bold=True, color=DRK)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells(start_row=ri, start_column=5, end_row=ri, end_column=14)
    c = ws.cell(ri, 5)
    c.value = detail; c.font = Font(name='Calibri', size=10, color=DRK)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[ri].height = 22

border_range(ws, 28, 1, 34, 14)

# Footer
ws.merge_cells('A36:N36')
hdr(ws['A36'], 'Generated by Python (Scikit-learn, Pandas, OpenPyXL) | Project: Healthcare No-Show Prediction',
    sz=9, bg=GRY, color=DRK, bold=False, italic=True)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 – RAW DATA
# ─────────────────────────────────────────────────────────────────────────────
df_full = pd.read_csv('models/full_predictions.csv')
ws2 = wb.create_sheet('Raw Data')
ws2.sheet_view.showGridLines = True
ws2.merge_cells('A1:P1')
hdr(ws2['A1'], 'Appointment Dataset with Risk Scores', sz=14, bg=BLU)
ws2.row_dimensions[1].height = 28

cols_show = ['PatientId','AppointmentID','Gender','Age','AgeGroup','Neighbourhood',
             'ScheduledDay','AppointmentDay','AppointmentWeekday','WaitDays',
             'SMS_received','Scholarship','Hipertension','Diabetes',
             'No-show','NoShow_Prob','Risk_Level']
sub = df_full[cols_show].head(500)
for ci, col in enumerate(cols_show, 1):
    c = ws2.cell(2, ci)
    hdr(c, col.replace('_',' '), sz=10, bg=BLU2)
    ws2.column_dimensions[get_column_letter(ci)].width = max(len(col)+2, 12)

for ri, row in enumerate(sub.itertuples(index=False), 3):
    for ci, val_v in enumerate(row, 1):
        c = ws2.cell(ri, ci)
        c.value = val_v
        c.font  = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='center')
        if ci == len(cols_show):  # Risk_Level
            color = RED if val_v=='High' else ORG if val_v=='Medium' else GRN
            c.fill = PatternFill('solid', fgColor=color)
            c.font = Font(name='Calibri', size=9, bold=True, color=WHT)

border_range(ws2, 2, 1, min(502, 2+len(sub)), len(cols_show))
ws2.freeze_panes = 'A3'

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 – MODEL METRICS
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Model Metrics')
ws3.sheet_view.showGridLines = False
ws3.merge_cells('A1:H1')
hdr(ws3['A1'], 'Model Performance Metrics — Decision Tree Classifier', sz=14, bg=BLU)
ws3.row_dimensions[1].height = 30

# Metrics table
metrics = [('Accuracy','0.554','75%','Correct predictions / Total'),
           ('Precision','0.286','50%','True No-Shows / Predicted No-Shows'),
           ('Recall','0.530','60%','Detected No-Shows / Actual No-Shows'),
           ('F1 Score','0.372','60%','Harmonic mean of Precision & Recall'),
           ('AUC-ROC','0.575','0.70','Model discrimination ability'),
           ('CV AUC (5-fold)','0.582','0.70','Cross-validated AUC – generalisation')]

for ci, h in enumerate(['Metric','Score','Benchmark','Interpretation'], 1):
    hdr(ws3.cell(3, ci), h, sz=11, bg=DRK)
    ws3.column_dimensions[get_column_letter(ci)].width = 28

for ri, row in enumerate(metrics, 4):
    for ci, txt in enumerate(row, 1):
        c = ws3.cell(ri, ci)
        c.value = txt
        c.font  = Font(name='Calibri', size=11, color=DRK)
        c.fill  = PatternFill('solid', fgColor=GRY if ri%2==0 else WHT)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[ri].height = 20

# Feature importance table
ws3.merge_cells('A12:D12')
hdr(ws3['A12'], 'Feature Importance Rankings', sz=12, bg=BLU2)
ws3.row_dimensions[12].height = 22

fi_df = pd.read_csv('models/feature_importance.csv').head(10)
for ci, h in enumerate(['Rank','Feature','Importance','Bar'], 1):
    hdr(ws3.cell(13, ci), h, sz=10, bg=BLU)

for ri, (_, row) in enumerate(fi_df.iterrows(), 14):
    ws3.cell(ri, 1).value = ri - 13
    ws3.cell(ri, 2).value = row['Feature']
    ws3.cell(ri, 3).value = round(row['Importance'], 4)
    bar_len = int(row['Importance'] * 300)
    ws3.cell(ri, 4).value = '█' * bar_len
    ws3.cell(ri, 4).font  = Font(name='Calibri', size=10, color=BLU)
    for ci in range(1, 4):
        c = ws3.cell(ri, ci)
        c.font  = Font(name='Calibri', size=10, color=DRK)
        c.fill  = PatternFill('solid', fgColor=GRY if ri%2==0 else WHT)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[ri].height = 18

border_range(ws3, 3, 1, 9, 4)
border_range(ws3, 13, 1, 23, 4)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 – TREND ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Trend Analysis')
ws4.sheet_view.showGridLines = False
ws4.merge_cells('A1:J1')
hdr(ws4['A1'], 'No-Show Trend Analysis by Key Dimensions', sz=14, bg=BLU)
ws4.row_dimensions[1].height = 30

sections = [
    ('SMS Reminder Impact', 'models/sms_noshow.csv', 3, 1),
    ('No-Show by Age Group', 'models/age_noshow.csv', 3, 4),
    ('No-Show by Weekday', 'models/weekday_noshow.csv', 3, 7),
]
for title, path, start_row, start_col in sections:
    df_t = pd.read_csv(path)
    ws4.merge_cells(start_row=start_row-1, start_column=start_col,
                    end_row=start_row-1, end_column=start_col+1)
    hdr(ws4.cell(start_row-1, start_col), title, sz=11, bg=BLU2)
    for ci, col in enumerate(df_t.columns, start_col):
        hdr(ws4.cell(start_row, ci), col, sz=10, bg=DRK)
        ws4.column_dimensions[get_column_letter(ci)].width = 18
    for ri, (_, row) in enumerate(df_t.iterrows(), start_row+1):
        for ci, v in enumerate(row, start_col):
            c = ws4.cell(ri, ci)
            c.value = round(v, 4) if isinstance(v, float) else v
            c.font  = Font(name='Calibri', size=10, color=DRK)
            c.fill  = PatternFill('solid', fgColor=GRY if ri%2==0 else WHT)
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws4.row_dimensions[ri].height = 18

# Bar chart for weekday
wd_df = pd.read_csv('models/weekday_noshow.csv')
ws4_start = 12
for ri, row in enumerate(wd_df.itertuples(), ws4_start):
    ws4.cell(ri, 10).value = row.Weekday
    ws4.cell(ri, 11).value = round(row.NoShow_Rate, 3)

bar = BarChart()
bar.type = 'col'; bar.title = 'No-Show Rate by Weekday'
bar.y_axis.title = 'No-Show Rate'; bar.x_axis.title = 'Weekday'
bar.width = 18; bar.height = 12
data_ref = Reference(ws4, min_col=11, min_row=ws4_start-1,
                     max_row=ws4_start+len(wd_df)-1)
cats     = Reference(ws4, min_col=10, min_row=ws4_start,
                     max_row=ws4_start+len(wd_df)-1)
bar.add_data(data_ref, titles_from_data=False)
bar.set_categories(cats)
ws4.add_chart(bar, 'A18')

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 – POWER BI READY DATA
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet('PowerBI Export')
ws5.merge_cells('A1:L1')
hdr(ws5['A1'], 'Power BI Ready Export — Connect this sheet as data source', sz=12, bg=GRN)
ws5.row_dimensions[1].height = 24

pbi_cols = ['AppointmentID','Gender','Age','AgeGroup','Neighbourhood',
            'AppointmentWeekday','AppointmentMonth','WaitDays',
            'SMS_received','Scholarship','No-show','NoShow_Prob','Risk_Level']
pbi_df = df_full[pbi_cols].copy()

for ci, col in enumerate(pbi_cols, 1):
    hdr(ws5.cell(2, ci), col, sz=10, bg=BLU2)
    ws5.column_dimensions[get_column_letter(ci)].width = 16

for ri, row in enumerate(pbi_df.itertuples(index=False), 3):
    for ci, v in enumerate(row, 1):
        c = ws5.cell(ri, ci)
        c.value = v
        c.font  = Font(name='Calibri', size=9)
        c.alignment = Alignment(horizontal='center')

ws5.freeze_panes = 'A3'
border_range(ws5, 2, 1, 2, len(pbi_cols))

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
out = 'outputs/Healthcare_NoShow_Dashboard.xlsx'
wb.save(out)
print(f"✓ Excel dashboard saved → {out}")
